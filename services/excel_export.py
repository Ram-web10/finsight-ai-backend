from openpyxl import Workbook
from openpyxl.styles import Font


def export_to_excel(result: dict, output_path: str):
    """
    Export complete financial analysis into an Excel workbook.
    """

    workbook = Workbook()

    # =====================================================
    # Sheet 1 - Company Information
    # =====================================================

    sheet = workbook.active
    sheet.title = "Company"

    sheet["A1"] = "Field"
    sheet["B1"] = "Value"

    sheet["A1"].font = Font(bold=True)
    sheet["B1"].font = Font(bold=True)

    company = result.get("company", {})

    row = 2

    for key, value in company.items():
        sheet.cell(row=row, column=1).value = key
        sheet.cell(row=row, column=2).value = str(value)
        row += 1

    # =====================================================
    # Sheet 2 - Financial Data
    # =====================================================

    financial_sheet = workbook.create_sheet("Financial Data")

    financial_sheet["A1"] = "Metric"
    financial_sheet["B1"] = "Value"

    financial_sheet["A1"].font = Font(bold=True)
    financial_sheet["B1"].font = Font(bold=True)

    row = 2

    for key, value in result.get("financial_data", {}).items():
        financial_sheet.cell(row=row, column=1).value = key
        financial_sheet.cell(row=row, column=2).value = value
        row += 1

    # =====================================================
    # Sheet 3 - Ratios
    # =====================================================

    ratio_sheet = workbook.create_sheet("Ratios")

    ratio_sheet["A1"] = "Ratio"
    ratio_sheet["B1"] = "Value"

    ratio_sheet["A1"].font = Font(bold=True)
    ratio_sheet["B1"].font = Font(bold=True)

    row = 2

    for key, value in result.get("ratios", {}).items():
        ratio_sheet.cell(row=row, column=1).value = key
        ratio_sheet.cell(row=row, column=2).value = value
        row += 1

    # =====================================================
    # Sheet 4 - SWOT
    # =====================================================

    swot_sheet = workbook.create_sheet("SWOT")

    swot_sheet["A1"] = "Category"
    swot_sheet["B1"] = "Points"

    swot_sheet["A1"].font = Font(bold=True)
    swot_sheet["B1"].font = Font(bold=True)

    row = 2

    swot = result.get("swot", {})

    for category, values in swot.items():

        swot_sheet.cell(row=row, column=1).value = category

        if isinstance(values, list):
            swot_sheet.cell(row=row, column=2).value = ", ".join(values)
        else:
            swot_sheet.cell(row=row, column=2).value = str(values)

        row += 1

    # =====================================================
    # Sheet 5 - AI Analysis
    # =====================================================

    ai_sheet = workbook.create_sheet("AI Analysis")

    ai_sheet["A1"] = "Analysis"

    ai_sheet["A1"].font = Font(bold=True)

    ai_sheet["A2"] = str(result.get("ai_analysis", ""))

    # =====================================================
    # Save
    # =====================================================

    workbook.save(output_path)

    return output_path